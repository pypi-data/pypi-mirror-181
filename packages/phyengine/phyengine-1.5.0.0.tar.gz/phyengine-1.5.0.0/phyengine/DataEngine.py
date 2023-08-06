import openpyxl as excel

LAST = -1
FIRST = 0

EXCEL = 'xslx'
TXT = 'txt'

class RecordableValue:
    def __init__(self, x_expression: str = "0", y_expression: str = "0", window = None, **kwargs):
        self.kwargs = kwargs
        def check_expr(self, expr: str):
            new_expr = expr
            for item in kwargs.keys():
                new_expr = new_expr.replace(item, "self.kwargs['{}']".format(item))
            try:
                eval(new_expr)
            except Exception as error:
                print(error)
                raise ValueError("can not calculate expression {}".format(new_expr))
            return new_expr
        self.data = list()
        self.x_expr = check_expr(self, x_expression)
        self.y_expr = check_expr(self, y_expression)
        self.window = window
        if window: self.window.InitObjects(self)

    def record(self):
        self.data.append([eval(self.x_expr), eval(self.y_expr)])

    def differentiate(self, index = LAST):
        if len(self.data) < 2:
            return 0
        if index == LAST:
            return (self.data[index][1] - self.data[index - 1][1]) / (self.data[index][0] - self.data[index - 1][0])
        elif index == FIRST:
            return (self.data[index + 1][1] - self.data[index][1]) / (self.data[index + 1][0] - self.data[index][0])
        else:
            return (self.data[index + 1][1] - self.data[index - 1][1]) / (self.data[index + 1][0] - self.data[index - 1][0])

    def save(self, name: str = "temp", file_type = TXT):
        if file_type == TXT:
            with open('{}.txt'.format(name), 'w') as file:
                file.write('\n'.join(map(lambda k: '{}!{}'.format(*k), self.data)))
        elif file_type == EXCEL:
            book = excel.Workbook()
            excel_data = book.create_sheet('Data', 0)
            for i in range(len(self.data)):
                cell1 = excel_data.cell(i + 1, 1)
                cell1.value = self.data[i][0]
                cell2 = excel_data.cell(i + 1, 2)
                cell2.value = self.data[i][1]
            book.save('{}.xlsx'.format(name))